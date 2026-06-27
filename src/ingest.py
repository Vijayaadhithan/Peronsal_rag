import argparse
import csv
import hashlib
import json
import logging
import re
import time
from datetime import date, datetime
from decimal import Decimal
from collections import Counter
from pathlib import Path
from typing import Any

import chromadb
import requests
from pypdf import PdfReader

from bm25_index import PersistentBM25Index
from settings import (
    CHROMA_DIR,
    CHUNK_OVERLAP,
    CHUNK_SIZE,
    COLLECTION_NAME,
    EMBED_MODEL,
    MYSQL_BM25_COLUMN,
    MYSQL_CONTENT_COLUMN,
    MYSQL_DATABASE,
    MYSQL_HOST,
    MYSQL_PASSWORD,
    MYSQL_PORT,
    MYSQL_TABLE,
    MYSQL_USER,
    MYSQL_SEARCH_ID_COLUMN,
    OLLAMA_BASE_URL,
    SOURCE_FILES_DIR,
)

EMBED_BATCH_SIZE = 32
MYSQL_BATCH_SIZE = 500
MYSQL_METADATA_MAX_CHARS = 1000
MYSQL_JSON_METADATA_MAX_FIELDS = 64
SUPPORTED_EXTENSIONS = {".pdf", ".csv", ".tsv", ".xlsx", ".xlsm"}
MYSQL_ID_CANDIDATES = ("id", "ad_id", "ads_id", "adId", "adsId")
JSON_TEXT_KEYS = (
    "embedding_text",
    "semantic_text",
    "search_text",
    "text",
    "content",
    "document",
)
LABELED_TEXT_KEYS = (
    "Title",
    "Description",
    "Listing meta title",
    "Listing meta description",
    "Main category",
    "Main category meta title",
    "Main category meta description",
    "Subcategory",
    "Subcategory meta title",
    "Subcategory meta description",
    "Listing rental duration",
    "State",
    "City",
    "Locality",
    "Selected attributes",
    "Selected attribute values",
)
LABELED_TEXT_PATTERN = re.compile(
    r"(?<!^)\s+("
    + "|".join(re.escape(label) for label in sorted(LABELED_TEXT_KEYS, key=len, reverse=True))
    + r"):\s*"
)


class _PypdfRepairFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        return not record.getMessage().startswith("Ignoring wrong pointing object")


logging.getLogger("pypdf._reader").addFilter(_PypdfRepairFilter())


def embed_texts(texts: list[str]) -> list[list[float]]:
    try:
        response = requests.post(
            f"{OLLAMA_BASE_URL}/api/embed",
            json={"model": EMBED_MODEL, "input": texts},
            timeout=300,
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise RuntimeError(
            f"Cannot get embeddings from Ollama at {OLLAMA_BASE_URL}. "
            f"Start Ollama and confirm '{EMBED_MODEL}' is installed."
        ) from exc

    embeddings = response.json().get("embeddings")
    if not embeddings or len(embeddings) != len(texts):
        raise RuntimeError("Ollama returned an invalid embedding response.")
    return embeddings


def read_pdf(path: Path) -> tuple[list[dict], int, int]:
    reader = PdfReader(path)
    pages = []
    empty_pages = 0

    for page_number, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            pages.append({"page": page_number, "text": text})
        else:
            empty_pages += 1

    return pages, empty_pages, len(reader.pages)


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_headers(values: list[Any]) -> list[str]:
    headers = []
    for index, value in enumerate(values, start=1):
        text = cell_to_text(value)
        headers.append(text or f"column_{index}")
    return headers


def row_to_text(headers: list[str], values: list[Any]) -> str:
    parts = []
    width = max(len(headers), len(values))
    for index in range(width):
        value = cell_to_text(values[index] if index < len(values) else "")
        if not value:
            continue
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        parts.append(f"{header}: {value}")
    return "; ".join(parts)


def read_delimited_file(path: Path) -> tuple[list[dict], int, int]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = []
    empty_rows = 0

    with path.open(newline="", encoding="utf-8-sig") as source_file:
        reader = csv.reader(source_file, delimiter=delimiter)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], 0, 0

        headers = normalize_headers(header_row)
        total_rows = 1
        for row_number, row in enumerate(reader, start=2):
            total_rows = row_number
            text = row_to_text(headers, row)
            if text:
                rows.append({"row": row_number, "text": text})
            else:
                empty_rows += 1

    return rows, empty_rows, total_rows


def read_excel_file(path: Path) -> tuple[list[dict], int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading Excel files requires openpyxl. Install requirements.txt first."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    empty_rows = 0
    total_rows = 0

    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            continue

        headers = normalize_headers(list(header_row))
        total_rows += 1
        for row_number, row in enumerate(iterator, start=2):
            total_rows += 1
            text = row_to_text(headers, list(row))
            if text:
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "row": row_number,
                        "text": text,
                    }
                )
            else:
                empty_rows += 1

    workbook.close()
    return rows, empty_rows, total_rows


def chunk_text(text: str, chunk_size: int = 512, overlap: int = 80) -> list[str]:
    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than zero")
    if overlap < 0 or overlap >= chunk_size:
        raise ValueError("overlap must be at least zero and smaller than chunk_size")

    words = text.split()
    step = chunk_size - overlap
    return [
        " ".join(words[start : start + chunk_size])
        for start in range(0, len(words), step)
    ]


def chunk_id(filename: str, location: int | str, index: int) -> str:
    value = f"{filename}\0{location}\0{index}".encode()
    return hashlib.sha256(value).hexdigest()


def mysql_source_name() -> str:
    return f"mysql:{MYSQL_DATABASE}.{MYSQL_TABLE}"


def mysql_document_id(table: str, row_identity: Any) -> str:
    value = f"mysql\0{MYSQL_DATABASE}\0{table}\0{row_identity}".encode()
    return hashlib.sha256(value).hexdigest()


def content_hash(document: str) -> str:
    return hashlib.sha256(document.encode()).hexdigest()


def quote_mysql_identifier(identifier: str) -> str:
    if not identifier or "\x00" in identifier:
        raise ValueError("MySQL identifiers must be non-empty strings")
    return f"`{identifier.replace('`', '``')}`"


def metadata_value(value: Any) -> bool | int | float | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        try:
            text = value.decode("utf-8")
        except UnicodeDecodeError:
            text = value.hex()
    else:
        text = str(value)
    text = text.strip()
    if not text:
        return None
    if len(text) > MYSQL_METADATA_MAX_CHARS:
        return text[:MYSQL_METADATA_MAX_CHARS]
    return text


def normalize_metadata_key(value: str) -> str:
    chars = []
    for char in value:
        chars.append(char if char.isalnum() else "_")
    normalized = "_".join(part for part in "".join(chars).split("_") if part)
    return normalized[:120] or "value"


def parse_json_like(value: Any) -> Any | None:
    if isinstance(value, (dict, list)):
        return value
    text = cell_to_text(value)
    if not text or text[0] not in "[{":
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def normalize_labeled_text(value: str) -> str:
    text = " ".join(value.split())
    return LABELED_TEXT_PATTERN.sub(r"\n\1: ", text).strip()


def extract_labeled_text_metadata(value: str) -> dict:
    matches = list(
        re.finditer(
            r"(?:^|\n)("
            + "|".join(
                re.escape(label)
                for label in sorted(LABELED_TEXT_KEYS, key=len, reverse=True)
            )
            + r"):\s*",
            value,
        )
    )
    metadata = {}
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        content = value[start:end].strip()
        if not content:
            continue
        key = f"content_{normalize_metadata_key(match.group(1).lower())}"
        safe_value = metadata_value(content)
        if safe_value is not None:
            metadata[key] = safe_value
    return metadata


def iter_json_scalars(value: Any, prefix: str = ""):
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from iter_json_scalars(child, child_prefix)
    elif isinstance(value, list):
        for index, child in enumerate(value, start=1):
            child_prefix = f"{prefix}.{index}" if prefix else str(index)
            yield from iter_json_scalars(child, child_prefix)
    else:
        safe_value = metadata_value(value)
        if safe_value is not None:
            yield prefix, safe_value


def collect_json_text_values(value: Any) -> list[str]:
    values = []
    if isinstance(value, dict):
        for key, child in value.items():
            key_text = str(key).lower()
            if key_text in JSON_TEXT_KEYS:
                if isinstance(child, (dict, list)):
                    values.extend(
                        str(safe_value)
                        for _, safe_value in iter_json_scalars(child)
                        if isinstance(safe_value, str) and safe_value
                    )
                else:
                    text = cell_to_text(child)
                    if text:
                        values.append(text)
            else:
                values.extend(collect_json_text_values(child))
    elif isinstance(value, list):
        for child in value:
            values.extend(collect_json_text_values(child))
    return values


def flatten_json_as_document(value: Any) -> str:
    parts = []
    for path, safe_value in iter_json_scalars(value):
        if path:
            parts.append(f"{path}: {safe_value}")
        else:
            parts.append(str(safe_value))
    return "; ".join(parts)


def prepare_content_document(value: Any) -> tuple[str, dict]:
    parsed = parse_json_like(value)
    if parsed is None:
        text = cell_to_text(value)
        if not text:
            return "", {"content_format": "text"}
        document = normalize_labeled_text(text)
        metadata = extract_labeled_text_metadata(document)
        metadata["content_format"] = "labeled_text" if metadata else "text"
        return document, metadata

    selected_texts = collect_json_text_values(parsed)
    if selected_texts:
        document = "\n".join(dict.fromkeys(selected_texts))
    else:
        document = flatten_json_as_document(parsed)

    metadata = {"content_format": "json"}
    for index, (path, safe_value) in enumerate(iter_json_scalars(parsed), start=1):
        if index > MYSQL_JSON_METADATA_MAX_FIELDS:
            metadata["content_metadata_truncated"] = True
            break
        key = f"content_{normalize_metadata_key(path)}"
        if key not in metadata:
            metadata[key] = safe_value
    return document, metadata


def mysql_row_identity(row: dict[str, Any], primary_key_column: str | None) -> str:
    if primary_key_column and row.get(primary_key_column) is not None:
        return str(row[primary_key_column])
    normalized = "|".join(
        f"{key}={metadata_value(row[key])}" for key in sorted(row.keys())
    )
    return hashlib.sha256(normalized.encode()).hexdigest()


def prepare_mysql_row(
    row: dict[str, Any],
    content_column: str,
    primary_key_column: str | None,
) -> tuple[str, str, dict] | None:
    document, content_metadata = prepare_content_document(row.get(content_column))
    if not document:
        return None

    identity = mysql_row_identity(row, primary_key_column)
    metadata = {
        "source_file": mysql_source_name(),
        "source_type": "mysql",
        "source_database": MYSQL_DATABASE,
        "source_table": MYSQL_TABLE,
        "embedding_model": EMBED_MODEL,
        "source_content_hash": content_hash(document),
    }
    metadata.update(content_metadata)
    if primary_key_column and row.get(primary_key_column) is not None:
        metadata["primary_key_column"] = primary_key_column
        metadata["primary_key_value"] = metadata_value(row[primary_key_column])

    for column, value in row.items():
        if column == content_column:
            continue
        safe_value = metadata_value(value)
        if safe_value is not None:
            metadata[column] = safe_value

    return mysql_document_id(MYSQL_TABLE, identity), document, metadata


def prepare_bm25_index_row(
    row: dict[str, Any],
    content_column: str,
    primary_key_column: str | None,
) -> dict | None:
    content = cell_to_text(row.get(content_column))
    if not content:
        return None

    identity = mysql_row_identity(row, primary_key_column)
    product_id = row.get(MYSQL_SEARCH_ID_COLUMN)
    if product_id is None:
        product_id = identity

    rental_fee = metadata_value(row.get("rental_fee"))
    if not isinstance(rental_fee, (int, float)):
        rental_fee = None

    return {
        "doc_id": mysql_document_id(MYSQL_TABLE, identity),
        "product_id": product_id,
        "content": content,
        "main_category_name": metadata_value(row.get("main_category_name")),
        "subcategory_name": metadata_value(row.get("subcategory_name")),
        "state_name": metadata_value(row.get("state_name")),
        "city_name": metadata_value(row.get("city_name")),
        "locality_name": metadata_value(row.get("locality_name")),
        "rental_duration": metadata_value(row.get("rental_duration")),
        "rental_fee": rental_fee,
    }


def require_pymysql():
    try:
        import pymysql
    except ImportError as exc:
        raise RuntimeError(
            "MySQL ingestion requires PyMySQL. Install requirements.txt first."
        ) from exc
    return pymysql


def mysql_connection(cursorclass=None):
    pymysql = require_pymysql()
    return pymysql.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DATABASE,
        charset="utf8mb4",
        autocommit=True,
        cursorclass=cursorclass,
        read_timeout=300,
        write_timeout=300,
    )


def fetch_mysql_columns() -> list[str]:
    pymysql = require_pymysql()
    with mysql_connection(cursorclass=pymysql.cursors.DictCursor) as connection:
        with connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM {quote_mysql_identifier(MYSQL_TABLE)}")
            return [row["Field"] for row in cursor.fetchall()]


def detect_mysql_primary_key(columns: list[str], override: str | None = None) -> str | None:
    if override:
        if override not in columns:
            raise RuntimeError(f"MySQL primary key column '{override}' was not found.")
        return override

    pymysql = require_pymysql()
    with mysql_connection(cursorclass=pymysql.cursors.DictCursor) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"SHOW KEYS FROM {quote_mysql_identifier(MYSQL_TABLE)} "
                "WHERE Key_name = 'PRIMARY'"
            )
            keys = cursor.fetchall()
            if keys:
                return keys[0]["Column_name"]

    for candidate in MYSQL_ID_CANDIDATES:
        if candidate in columns:
            return candidate
    return None


def count_mysql_rows(content_column: str) -> int:
    pymysql = require_pymysql()
    where_clause = (
        f"{quote_mysql_identifier(content_column)} IS NOT NULL "
        f"AND TRIM({quote_mysql_identifier(content_column)}) <> ''"
    )
    with mysql_connection(cursorclass=pymysql.cursors.DictCursor) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT COUNT(*) AS row_count FROM {quote_mysql_identifier(MYSQL_TABLE)} "
                f"WHERE {where_clause}"
            )
            return int(cursor.fetchone()["row_count"])


def iter_mysql_rows(
    content_column: str,
    primary_key_column: str | None,
    limit: int | None = None,
):
    pymysql = require_pymysql()
    quoted_content = quote_mysql_identifier(content_column)
    query = (
        f"SELECT * FROM {quote_mysql_identifier(MYSQL_TABLE)} "
        f"WHERE {quoted_content} IS NOT NULL AND TRIM({quoted_content}) <> ''"
    )
    params = []
    if primary_key_column:
        query += f" ORDER BY {quote_mysql_identifier(primary_key_column)}"
    if limit is not None:
        query += " LIMIT %s"
        params.append(limit)

    with mysql_connection(cursorclass=pymysql.cursors.SSDictCursor) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            for row in cursor:
                yield row


def prepare_pdf(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    pages, empty_pages, page_count = read_pdf(path)
    ids = []
    documents = []
    metadatas = []

    for page in pages:
        for index, text in enumerate(
            chunk_text(page["text"], CHUNK_SIZE, CHUNK_OVERLAP)
        ):
            ids.append(chunk_id(path.name, page["page"], index))
            documents.append(text)
            metadatas.append(
                {
                    "source_file": path.name,
                    "page": page["page"],
                    "chunk_index": index,
                    "embedding_model": EMBED_MODEL,
                }
            )

    return ids, documents, metadatas, empty_pages, page_count


def prepare_table(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    if path.suffix.lower() in {".csv", ".tsv"}:
        rows, empty_rows, row_count = read_delimited_file(path)
    else:
        rows, empty_rows, row_count = read_excel_file(path)

    ids = []
    documents = []
    metadatas = []

    for row in rows:
        location = f"{row['sheet']}:{row['row']}" if "sheet" in row else str(row["row"])
        for index, text in enumerate(chunk_text(row["text"], CHUNK_SIZE, CHUNK_OVERLAP)):
            ids.append(chunk_id(path.name, location, index))
            documents.append(text)
            metadata = {
                "source_file": path.name,
                "source_type": path.suffix.lower().lstrip("."),
                "row": row["row"],
                "chunk_index": index,
                "embedding_model": EMBED_MODEL,
            }
            if "sheet" in row:
                metadata["sheet"] = row["sheet"]
            metadatas.append(metadata)

    return ids, documents, metadatas, empty_rows, row_count


def prepare_source(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    if path.suffix.lower() == ".pdf":
        return prepare_pdf(path)
    if path.suffix.lower() in SUPPORTED_EXTENSIONS:
        return prepare_table(path)
    raise ValueError(f"Unsupported source type: {path.suffix}")


def find_source_files() -> list[Path]:
    return sorted(
        [
            path
            for path in SOURCE_FILES_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
        ],
        key=lambda path: path.name.lower(),
    )


def get_collection(create: bool = False):
    client = chromadb.PersistentClient(path=str(CHROMA_DIR))
    if create:
        collection = client.get_or_create_collection(
            name=COLLECTION_NAME,
            metadata={"hnsw:space": "cosine"},
        )
        return client, collection
    try:
        return client, client.get_collection(COLLECTION_NAME)
    except Exception as exc:
        raise RuntimeError("No vector collection found. Run: python src/ingest.py") from exc


def list_indexed_documents() -> None:
    _, collection = get_collection()
    data = collection.get(include=["metadatas"])
    counts = Counter(metadata["source_file"] for metadata in data["metadatas"])

    print(f"Collection: {COLLECTION_NAME} ({collection.count()} chunks)\n")
    for filename, count in sorted(counts.items(), key=lambda item: item[0].lower()):
        print(f"{count:>5} chunks  {filename}")


def confirm(message: str, assume_yes: bool = False) -> bool:
    return assume_yes or input(f"{message} [y/N]: ").strip().lower() in {"y", "yes"}


def delete_indexed_document(filename: str, assume_yes: bool = False) -> None:
    _, collection = get_collection()
    existing = collection.get(where={"source_file": filename}, include=[])
    count = len(existing["ids"])
    if not count:
        raise RuntimeError(f"'{filename}' is not present in the collection.")
    if not confirm(
        f"Delete {count} indexed chunks for '{filename}'? The source file will be kept.",
        assume_yes,
    ):
        print("Cancelled.")
        return
    collection.delete(where={"source_file": filename})
    print(f"Deleted {count} chunks for '{filename}'.")


def clear_collection(assume_yes: bool = False) -> None:
    client, collection = get_collection()
    count = collection.count()
    if not confirm(
        f"Delete the entire '{COLLECTION_NAME}' collection ({count} chunks)?",
        assume_yes,
    ):
        print("Cancelled.")
        return
    client.delete_collection(COLLECTION_NAME)
    print(f"Deleted collection '{COLLECTION_NAME}'. Source files were kept.")


def source_is_current(
    collection, filename: str, ids: list[str], documents: list[str]
) -> bool:
    existing = collection.get(
        where={"source_file": filename},
        include=["documents", "metadatas"],
    )
    if len(existing["ids"]) != len(ids):
        return False

    stored_documents = dict(zip(existing["ids"], existing["documents"]))
    expected_documents = dict(zip(ids, documents))
    models_match = all(
        metadata.get("embedding_model") == EMBED_MODEL
        for metadata in existing["metadatas"]
    )
    return models_match and stored_documents == expected_documents


def check_sources(source_files: list[Path]) -> bool:
    valid = True
    total_units = 0
    total_chunks = 0
    for path in source_files:
        try:
            ids, _, _, skipped_units, unit_count = prepare_source(path)
            total_units += unit_count
            total_chunks += len(ids)
            print(
                f"OK: {path.name} | {unit_count} source units | {len(ids)} chunks | "
                f"{skipped_units} empty units"
            )
        except Exception as exc:
            valid = False
            print(f"ERROR: {path.name} | {type(exc).__name__}: {exc}")

    print(
        f"\nChecked {len(source_files)} source files, "
        f"{total_units} source units, {total_chunks} chunks."
    )
    return valid


def check_mysql_source(
    limit: int | None = None,
    primary_key_column: str | None = None,
) -> bool:
    columns = fetch_mysql_columns()
    if MYSQL_CONTENT_COLUMN not in columns:
        print(
            f"ERROR: column '{MYSQL_CONTENT_COLUMN}' was not found in "
            f"{MYSQL_DATABASE}.{MYSQL_TABLE}."
        )
        print(f"Available columns: {', '.join(columns)}")
        return False

    detected_primary_key = detect_mysql_primary_key(columns, primary_key_column)
    row_count = count_mysql_rows(MYSQL_CONTENT_COLUMN)
    planned_rows = min(row_count, limit) if limit is not None else row_count

    print(f"OK: MySQL table {MYSQL_DATABASE}.{MYSQL_TABLE}")
    print(f"Content column: {MYSQL_CONTENT_COLUMN}")
    print(f"Primary key column: {detected_primary_key or 'none detected'}")
    print(f"Rows with embedding text: {row_count}")
    print(f"Rows planned for ingestion: {planned_rows}")
    print("No embeddings were generated during this check.")
    return True


def embed_for_upsert(
    documents: list[str],
    embed_batch_size: int = EMBED_BATCH_SIZE,
    progress_prefix: str = "",
) -> list[list[float]]:
    embeddings = []
    for start in range(0, len(documents), embed_batch_size):
        batch = documents[start : start + embed_batch_size]
        if progress_prefix:
            completed = min(start + len(batch), len(documents))
            print(
                f"{progress_prefix} embedding {completed}/{len(documents)} texts",
                flush=True,
            )
        embeddings.extend(embed_texts(batch))
    return embeddings


def mysql_current_ids(
    collection,
    ids: list[str],
    documents: list[str],
    metadatas: list[dict],
) -> set[str]:
    if not ids:
        return set()
    existing = collection.get(ids=ids, include=["documents", "metadatas"])
    expected = {
        doc_id: {
            "document": document,
            "hash": metadata.get("source_content_hash"),
        }
        for doc_id, document, metadata in zip(ids, documents, metadatas)
    }
    current = set()
    for doc_id, document, metadata in zip(
        existing["ids"], existing["documents"], existing["metadatas"]
    ):
        if metadata.get("embedding_model") != EMBED_MODEL:
            continue
        expected_row = expected.get(doc_id, {})
        hash_matches = metadata.get("source_content_hash") == expected_row.get("hash")
        document_matches = document == expected_row.get("document")
        if hash_matches or document_matches:
            current.add(doc_id)
    return current


def format_duration(seconds: float) -> str:
    if seconds < 60:
        return f"{seconds:.0f}s"
    minutes, secs = divmod(int(seconds), 60)
    if minutes < 60:
        return f"{minutes}m {secs}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes}m"


def ingest_sources(source_files: list[Path]) -> None:
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    _, collection = get_collection(create=True)

    for path in source_files:
        print(f"Processing: {path.name}")
        try:
            ids, documents, metadatas, skipped_units, _ = prepare_source(path)
            if not documents:
                print("  Skipped: no extractable text (OCR may be required).")
                continue
            if source_is_current(collection, path.name, ids, documents):
                print(f"  Unchanged: keeping {len(documents)} existing chunks.")
                continue

            embeddings = []
            for start in range(0, len(documents), EMBED_BATCH_SIZE):
                batch = documents[start : start + EMBED_BATCH_SIZE]
                embeddings.extend(embed_texts(batch))
                completed = min(start + len(batch), len(documents))
                print(
                    f"  Embedded {completed}/{len(documents)} chunks",
                    end="\r",
                    flush=True,
                )
            print()

            # Remove stale chunks only after extraction and embedding have succeeded.
            collection.delete(where={"source_file": path.name})
            collection.upsert(
                ids=ids,
                documents=documents,
                embeddings=embeddings,
                metadatas=metadatas,
            )
            print(f"  Added {len(documents)} chunks; skipped {skipped_units} empty units.")
        except RuntimeError:
            raise
        except Exception as exc:
            print(f"  ERROR: {type(exc).__name__}: {exc}")

    print(f"\nIngestion complete. Collection contains {collection.count()} chunks.")


def ingest_mysql_source(
    limit: int | None = None,
    batch_size: int = MYSQL_BATCH_SIZE,
    embed_batch_size: int = EMBED_BATCH_SIZE,
    primary_key_column: str | None = None,
    replace_source: bool = False,
    force_reembed: bool = False,
) -> None:
    if batch_size <= 0:
        raise RuntimeError("--mysql-batch-size must be greater than zero.")
    if embed_batch_size <= 0:
        raise RuntimeError("--embed-batch-size must be greater than zero.")
    if limit is not None and limit <= 0:
        raise RuntimeError("--limit must be greater than zero.")

    columns = fetch_mysql_columns()
    if MYSQL_CONTENT_COLUMN not in columns:
        raise RuntimeError(
            f"Column '{MYSQL_CONTENT_COLUMN}' was not found in "
            f"{MYSQL_DATABASE}.{MYSQL_TABLE}."
        )
    detected_primary_key = detect_mysql_primary_key(columns, primary_key_column)
    bm25_column = (
        MYSQL_BM25_COLUMN
        if MYSQL_BM25_COLUMN in columns
        else MYSQL_CONTENT_COLUMN
    )
    row_count = count_mysql_rows(MYSQL_CONTENT_COLUMN)
    planned_rows = min(row_count, limit) if limit is not None else row_count

    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    _, collection = get_collection(create=True)
    bm25_index = PersistentBM25Index()
    source_name = mysql_source_name()

    print(f"Processing MySQL table: {MYSQL_DATABASE}.{MYSQL_TABLE}")
    print(f"Content column: {MYSQL_CONTENT_COLUMN}")
    print(f"BM25 column: {bm25_column}")
    print(f"Primary key column: {detected_primary_key or 'none detected'}")
    print(f"Rows planned for ingestion: {planned_rows}")

    if replace_source:
        existing = collection.get(where={"source_file": source_name}, include=[])
        if existing["ids"]:
            collection.delete(where={"source_file": source_name})
            print(f"Deleted {len(existing['ids'])} existing chunks for {source_name}.")
        bm25_index.clear()
        print("Cleared the persistent BM25 product index.")

    ids: list[str] = []
    documents: list[str] = []
    metadatas: list[dict] = []
    bm25_rows: list[dict] = []
    indexed = 0
    processed = 0
    skipped_empty = 0
    skipped_current = 0
    started_at = time.monotonic()

    def flush_batch() -> None:
        nonlocal ids, documents, metadatas, bm25_rows, indexed, skipped_current
        if not documents:
            return
        batch_start = processed - len(documents) + 1
        batch_end = processed
        total_label = planned_rows if planned_rows else "unknown"
        bm25_index.upsert(bm25_rows)

        if force_reembed:
            upsert_ids = ids
            upsert_documents = documents
            upsert_metadatas = metadatas
        else:
            current_ids = mysql_current_ids(collection, ids, documents, metadatas)
            skipped_current += len(current_ids)
            upsert_ids = []
            upsert_documents = []
            upsert_metadatas = []
            for doc_id, document, metadata in zip(ids, documents, metadatas):
                if doc_id in current_ids:
                    continue
                upsert_ids.append(doc_id)
                upsert_documents.append(document)
                upsert_metadatas.append(metadata)

        if not upsert_documents:
            print(
                f"  Rows {batch_start}-{batch_end}/{total_label} unchanged; skipped.",
                flush=True,
            )
            ids = []
            documents = []
            metadatas = []
            bm25_rows = []
            return

        print(
            f"  Preparing rows {batch_start}-{batch_end}/{total_label} for Chroma "
            f"({len(upsert_documents)} changed/new)",
            flush=True,
        )
        embeddings = embed_for_upsert(
            upsert_documents,
            embed_batch_size,
            progress_prefix=f"    rows {batch_start}-{batch_end}",
        )
        collection.upsert(
            ids=upsert_ids,
            documents=upsert_documents,
            embeddings=embeddings,
            metadatas=upsert_metadatas,
        )
        indexed += len(upsert_documents)
        elapsed = time.monotonic() - started_at
        rate = processed / elapsed if elapsed else 0
        remaining = max(planned_rows - processed, 0)
        eta = remaining / rate if rate else 0
        print(
            f"  Indexed/updated {indexed} rows; skipped unchanged {skipped_current}; "
            f"processed {processed}/{total_label}; ETA {format_duration(eta)}",
            flush=True,
        )
        ids = []
        documents = []
        metadatas = []
        bm25_rows = []

    for row in iter_mysql_rows(MYSQL_CONTENT_COLUMN, detected_primary_key, limit):
        prepared = prepare_mysql_row(row, MYSQL_CONTENT_COLUMN, detected_primary_key)
        if prepared is None:
            skipped_empty += 1
            continue

        doc_id, document, metadata = prepared
        ids.append(doc_id)
        documents.append(document)
        metadatas.append(metadata)
        bm25_row = prepare_bm25_index_row(
            row,
            bm25_column,
            detected_primary_key,
        )
        if bm25_row is not None:
            bm25_rows.append(bm25_row)
        processed += 1

        if len(documents) >= batch_size:
            flush_batch()

    flush_batch()
    bm25_count = bm25_index.count()
    bm25_index.close()
    print(
        f"\nMySQL ingestion complete. Indexed/updated {indexed} rows; "
        f"skipped unchanged {skipped_current} rows; skipped empty {skipped_empty} rows. "
        f"Collection contains {collection.count()} chunks. "
        f"BM25 index contains {bm25_count} products."
    )


def rebuild_mysql_bm25_index(
    limit: int | None = None,
    batch_size: int = MYSQL_BATCH_SIZE,
    primary_key_column: str | None = None,
) -> None:
    if batch_size <= 0:
        raise RuntimeError("--mysql-batch-size must be greater than zero.")
    if limit is not None and limit <= 0:
        raise RuntimeError("--limit must be greater than zero.")

    columns = fetch_mysql_columns()
    if MYSQL_CONTENT_COLUMN not in columns:
        raise RuntimeError(
            f"Column '{MYSQL_CONTENT_COLUMN}' was not found in "
            f"{MYSQL_DATABASE}.{MYSQL_TABLE}."
        )
    detected_primary_key = detect_mysql_primary_key(columns, primary_key_column)
    bm25_column = (
        MYSQL_BM25_COLUMN
        if MYSQL_BM25_COLUMN in columns
        else MYSQL_CONTENT_COLUMN
    )
    row_count = count_mysql_rows(MYSQL_CONTENT_COLUMN)
    planned_rows = min(row_count, limit) if limit is not None else row_count

    index = PersistentBM25Index()
    index.clear()
    batch = []
    processed = 0

    print(f"Rebuilding BM25 index from {MYSQL_DATABASE}.{MYSQL_TABLE}")
    print(f"BM25 column: {bm25_column}")
    print(f"Rows planned: {planned_rows}")

    for row in iter_mysql_rows(MYSQL_CONTENT_COLUMN, detected_primary_key, limit):
        entry = prepare_bm25_index_row(
            row,
            bm25_column,
            detected_primary_key,
        )
        if entry is None:
            continue
        batch.append(entry)
        processed += 1
        if len(batch) >= batch_size:
            index.upsert(batch)
            batch = []
            print(f"  Indexed {processed}/{planned_rows}", end="\r", flush=True)

    index.upsert(batch)
    count = index.count()
    index.close()
    print(f"\nBM25 rebuild complete. Indexed {count} products.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest local source files into Chroma.")
    actions = parser.add_mutually_exclusive_group()
    actions.add_argument(
        "--check", action="store_true", help="validate source files without calling Ollama"
    )
    actions.add_argument(
        "--list", action="store_true", help="list indexed source files and chunk counts"
    )
    actions.add_argument(
        "--delete", metavar="FILENAME", help="delete one source file from the index"
    )
    actions.add_argument(
        "--clear", action="store_true", help="delete the entire vector collection"
    )
    parser.add_argument(
        "--yes", action="store_true", help="skip confirmation for delete operations"
    )
    parser.add_argument(
        "--mysql",
        action="store_true",
        help="ingest rows from the configured MySQL table instead of local files",
    )
    parser.add_argument(
        "--mysql-bm25-only",
        action="store_true",
        help="rebuild only the persistent BM25 index from MySQL; no embeddings",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="limit rows for MySQL smoke-test ingestion",
    )
    parser.add_argument(
        "--mysql-batch-size",
        type=int,
        default=MYSQL_BATCH_SIZE,
        help=f"MySQL rows to upsert per Chroma batch (default: {MYSQL_BATCH_SIZE})",
    )
    parser.add_argument(
        "--embed-batch-size",
        type=int,
        default=EMBED_BATCH_SIZE,
        help=f"texts to send per Ollama embedding request (default: {EMBED_BATCH_SIZE})",
    )
    parser.add_argument(
        "--mysql-primary-key",
        help="override the detected MySQL primary key column",
    )
    parser.add_argument(
        "--mysql-replace-source",
        action="store_true",
        help="delete existing Chroma chunks for this MySQL source before ingesting",
    )
    parser.add_argument(
        "--mysql-force-reembed",
        action="store_true",
        help="re-embed MySQL rows even when the existing Chroma content hash matches",
    )
    args = parser.parse_args()

    if args.list:
        list_indexed_documents()
        return
    if args.delete:
        delete_indexed_document(args.delete, args.yes)
        return
    if args.clear:
        clear_collection(args.yes)
        return
    if args.mysql and args.check:
        raise SystemExit(
            0
            if check_mysql_source(args.limit, args.mysql_primary_key)
            else 1
        )
    if args.mysql_bm25_only:
        rebuild_mysql_bm25_index(
            limit=args.limit,
            batch_size=args.mysql_batch_size,
            primary_key_column=args.mysql_primary_key,
        )
        return
    if args.mysql:
        ingest_mysql_source(
            limit=args.limit,
            batch_size=args.mysql_batch_size,
            embed_batch_size=args.embed_batch_size,
            primary_key_column=args.mysql_primary_key,
            replace_source=args.mysql_replace_source,
            force_reembed=args.mysql_force_reembed,
        )
        return

    source_files = find_source_files()
    if not source_files:
        extensions = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise SystemExit(f"No supported files ({extensions}) found in {SOURCE_FILES_DIR}")

    if args.check:
        raise SystemExit(0 if check_sources(source_files) else 1)
    ingest_sources(source_files)


if __name__ == "__main__":
    try:
        main()
    except RuntimeError as exc:
        raise SystemExit(f"Error: {exc}") from exc
